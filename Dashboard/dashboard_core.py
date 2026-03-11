"""
================================================================================
  FILE: dashboard_core.py
  LOCATION: FnOBot/Dashboard/dashboard_core.py

  CHANGES (Phase 1 Rank 1):
    - DATA_FOLDER fixed: C:/Users/marut/Desktop/Market Data (forward slashes)
    - DASH_FILE written to DATA_FOLDER
    - Dual-expiry tabs: Current Expiry + Next Expiry in option chain table
    - OI delta columns: 3-min, 1-hr, and Previous Day
    - Dashboard auto-refresh stops at 15:45 IST via JS time check
    - India VIX displayed in header bar
    - EOD: archives dashboard to DashboardArchive/dashboard_YYYY-MM-DD.html
================================================================================
"""

import os
import sys
import shutil
import datetime
import openpyxl
from zoneinfo import ZoneInfo

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

try:
    from indicators import build_vwap_html, get_vwap_css
    VWAP_AVAILABLE = True
except ImportError:
    VWAP_AVAILABLE = False
    def build_vwap_html(vwap_data, symbol): return ""
    def get_vwap_css(): return ""

try:
    from indicators import build_technicals_html, get_technicals_css
    TECHNICALS_AVAILABLE = True
except ImportError:
    TECHNICALS_AVAILABLE = False
    def build_technicals_html(tech_data, symbol): return ""
    def get_technicals_css(): return ""

try:
    from indicators import build_iv_html, get_iv_css
    IV_AVAILABLE = True
except ImportError:
    IV_AVAILABLE = False
    def build_iv_html(iv_data, symbol): return ""
    def get_iv_css(): return ""

try:
    from indicators import build_oi_concentration_html, get_oi_concentration_css
    CONC_AVAILABLE = True
except ImportError:
    CONC_AVAILABLE = False
    def build_oi_concentration_html(conc_data, symbol): return ""
    def get_oi_concentration_css(): return ""

IST            = ZoneInfo("Asia/Kolkata")
DATA_FOLDER    = "C:/Users/marut/Desktop/Market Data"
ARCHIVE_FOLDER = "C:/Users/marut/Desktop/Market Data/DashboardArchive"
DASH_FILE      = "C:/Users/marut/Desktop/Market Data/dashboard.html"


# =============================================================================
#  EXCEL OI HISTORY READER
# =============================================================================

def _get_daily_filepath():
    today = datetime.date.today().strftime("%Y-%m-%d")
    return os.path.join(DATA_FOLDER, f"MarketData_{today}.xlsx")


def load_oi_from_excel(symbol: str, minutes_ago: int, tolerance_min: int = 45) -> dict:
    """Load OI snapshot from Excel for X minutes ago (used for 1-hr and prev-day deltas)."""
    try:
        filepath   = _get_daily_filepath()
        sheet_name = f"{symbol}_OI"
        if not os.path.exists(filepath):
            return {}
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {}
        ws        = wb[sheet_name]
        all_rows  = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(all_rows) < 2:
            return {}
        hdr       = list(all_rows[0])
        data_rows = all_rows[1:]
        try:
            ts_col     = hdr.index("Timestamp")
            strike_col = hdr.index("Strike")
            ce_col     = hdr.index("CE_OI")
            pe_col     = hdr.index("PE_OI")
        except ValueError:
            return {}
        now_ist    = datetime.datetime.now(IST)
        target_ist = now_ist - datetime.timedelta(minutes=minutes_ago)
        from collections import defaultdict
        ts_groups = defaultdict(list)
        for row in data_rows:
            ts_val = row[ts_col]
            if ts_val is not None:
                ts_groups[str(ts_val)].append(row)
        if not ts_groups:
            return {}

        def parse_ts(ts_str):
            for fmt in ["%H:%M", "%d-%b-%Y %H:%M", "%Y-%m-%d %H:%M:%S"]:
                try:
                    t = datetime.datetime.strptime(ts_str.strip(), fmt)
                    if t.year == 1900:
                        t = t.replace(year=now_ist.year, month=now_ist.month, day=now_ist.day)
                    return t.replace(tzinfo=IST)
                except Exception:
                    continue
            return None

        best_ts = best_diff = None
        for ts_str in ts_groups:
            t = parse_ts(ts_str)
            if t is None:
                continue
            diff = abs((t - target_ist).total_seconds())
            if best_diff is None or diff < best_diff:
                best_diff = diff
                best_ts   = ts_str
        if best_ts is None or best_diff > tolerance_min * 60:
            return {}
        snapshot = {}
        for row in ts_groups[best_ts]:
            strike = row[strike_col]
            if strike is not None:
                snapshot[float(strike)] = {
                    "CE": int(row[ce_col] or 0),
                    "PE": int(row[pe_col] or 0),
                }
        return snapshot
    except Exception:
        return {}


# =============================================================================
#  HELPERS
# =============================================================================

def _sig_style(sig):
    if "STRONG BULL" in sig:  return "#2ecc71", "#0d1f14", "#1a4a2a"
    if "MILD BULL"   in sig:  return "#52be80", "#0d1a10", "#1a3a20"
    if "NEUTRAL"     in sig:  return "#e67e22", "#1f1500", "#3a2800"
    if "MILD BEAR"   in sig:  return "#e74c3c", "#1f0a0a", "#4a1a1a"
    if "STRONG BEAR" in sig:  return "#c0392b", "#1a0808", "#3a1010"
    return "#94a3b8", "#0f172a", "#1e293b"


def _fmt_oi(v):
    if v is None:
        return "---"
    v = int(v)
    if abs(v) >= 1_000_000:
        return f"{v/1_000_000:.2f}M"
    if abs(v) >= 1_000:
        return f"{v/1_000:.1f}K"
    return str(v)


def _delta_span(val):
    if val is None:
        return '<span style="color:#4a5568">---</span>'
    try:
        n   = int(val)
        col = "#69f0ae" if n > 0 else "#ff6e6e" if n < 0 else "#4a5568"
        txt = f"+{_fmt_oi(n)}" if n > 0 else _fmt_oi(n)
        return f'<span style="color:{col};font-weight:600">{txt}</span>'
    except Exception:
        return '<span style="color:#4a5568">---</span>'


# =============================================================================
#  OPTION CHAIN TABLE BUILDER
# =============================================================================

def _build_chain_table(symbol: str, chain: dict, analysis: dict,
                       delta_data: dict, expiry_label: str,
                       oi_1hr: dict = None, oi_prevday: dict = None) -> str:
    if not chain:
        return f'<div style="color:#64748b;padding:20px">No chain data for {expiry_label}</div>'

    spot    = analysis.get("spot", 0) or 0
    atm     = analysis.get("atm", 0)
    strikes = sorted(chain.keys())
    deltas  = {d["Strike"]: d for d in delta_data.get("deltas", [])} if delta_data else {}

    col_hdr = (
        '<tr class="ch">'
        '<th class="ce-head" colspan="4">CALLS</th>'
        '<th class="strike-head">STRIKE</th>'
        '<th class="pe-head" colspan="4">PUTS</th>'
        '</tr>'
        '<tr class="ch2">'
        '<th class="ce-head">OI</th>'
        '<th class="ce-head">3m</th>'
        '<th class="ce-head">1h</th>'
        '<th class="ce-head">PD</th>'
        '<th class="strike-head">---</th>'
        '<th class="pe-head">OI</th>'
        '<th class="pe-head">3m</th>'
        '<th class="pe-head">1h</th>'
        '<th class="pe-head">PD</th>'
        '</tr>'
    )

    idx = strikes.index(atm) if atm in strikes else len(strikes) // 2
    window = strikes[max(0, idx - 5): idx + 6]
    rows = []
    for s in window:
        ce_oi = chain[s].get("CE", {}).get("openInterest", 0)
        pe_oi = chain[s].get("PE", {}).get("openInterest", 0)
        dk    = deltas.get(s, {})

        ce_3m = dk.get("CE_delta")
        pe_3m = dk.get("PE_delta")

        ce_1h = pe_1h = None
        if oi_1hr:
            prev1h = oi_1hr.get(s, {})
            ce_1h  = ce_oi - prev1h.get("CE", ce_oi)
            pe_1h  = pe_oi - prev1h.get("PE", pe_oi)

        ce_pd = pe_pd = None
        if oi_prevday:
            prev_pd = oi_prevday.get(s, {})
            ce_pd   = ce_oi - prev_pd.get("CE", ce_oi)
            pe_pd   = pe_oi - prev_pd.get("PE", pe_oi)

        atm_cls      = ' class="atm-row"' if s == atm else ""
        strike_label = f"{int(s):,}" + (" <" if s == atm else "")

        rows.append(
            f'<tr{atm_cls}>'
            f'<td class="r oi-num">{_fmt_oi(ce_oi)}</td>'
            f'<td class="r d15">{_delta_span(ce_3m)}</td>'
            f'<td class="r d1h">{_delta_span(ce_1h)}</td>'
            f'<td class="r dpd">{_delta_span(ce_pd)}</td>'
            f'<td class="strike-col">{strike_label}</td>'
            f'<td class="oi-num">{_fmt_oi(pe_oi)}</td>'
            f'<td class="d15">{_delta_span(pe_3m)}</td>'
            f'<td class="d1h">{_delta_span(pe_1h)}</td>'
            f'<td class="dpd">{_delta_span(pe_pd)}</td>'
            f'</tr>'
        )

    return (
        f'<table class="ct"><thead>{col_hdr}</thead>'
        f'<tbody>{"".join(rows)}</tbody></table>'
        f'<div class="chain-legend">'
        f'<span>3m = 3-min OI delta</span>'
        f'<span>1h = 1-hour OI delta</span>'
        f'<span>PD = vs previous day</span>'
        f'<span>Yellow = ATM</span>'
        f'</div>'
    )


# =============================================================================
#  MAIN DASHBOARD GENERATOR
# =============================================================================

def generate_dashboard(
    oi_data_map:    dict,
    analysis_map:   dict,
    delta_map:      dict,
    signals_map:    dict,
    vwap_map:       dict,
    tech_map:       dict,
    iv_map:         dict,
    conc_map:       dict,
    setup_map:      dict,
    futures_map:    dict,
    sr_map:         dict,
    nifty50_stocks: list  = None,
    india_vix:      float = 0.0,
    next_expiry_oi: dict  = None,
    next_analysis:  dict  = None,
    scan_count:     int   = 0,
    last_excel_ts:  str   = "",
    **kwargs,
) -> str:

    now_ist  = datetime.datetime.now(IST)
    ts_str   = now_ist.strftime("%d %b %Y  %H:%M:%S IST")
    stop_h, stop_m = 15, 45

    oi_1hr_map     = {}
    oi_prevday_map = {}
    for sym in oi_data_map:
        oi_1hr_map[sym]     = load_oi_from_excel(sym, 60)
        oi_prevday_map[sym] = load_oi_from_excel(sym, 24 * 60, tolerance_min=8 * 60)

    # ── Build per-symbol tab content ────────────────────────────────────────
    tab_contents = {}
    for sym in ["NIFTY", "BANKNIFTY"]:
        oi_data  = oi_data_map.get(sym, {})
        analysis = analysis_map.get(sym, {})
        delta    = delta_map.get(sym, {})
        signal, reasons = signals_map.get(sym, ("NEUTRAL", []))
        vwap     = vwap_map.get(sym, {})
        td       = tech_map.get(sym, {})
        iv       = iv_map.get(sym, {})
        conc     = conc_map.get(sym, {})
        setup    = setup_map.get(sym, {})
        fut      = futures_map.get(sym, {})
        next_oi  = (next_expiry_oi or {}).get(sym, {})
        next_ana = (next_analysis or {}).get(sym, {})

        spot     = oi_data.get("spot", 0)
        pcr      = analysis.get("pcr", 0)
        max_pain = analysis.get("max_pain", 0)
        expiry   = oi_data.get("expiry", "---")
        next_exp = next_oi.get("expiry", "---") if next_oi else "---"

        sig_col, sig_bg, sig_border = _sig_style(signal)
        best_setup  = setup.get("best_label", "---")
        best_score  = setup.get("best_score", 0)
        mood        = delta.get("mood", "---")
        reasons_str = "  .  ".join(reasons[:3]) if reasons else ""

        curr_table = _build_chain_table(
            sym, oi_data.get("chain", {}), {**analysis, "spot": spot},
            delta, expiry, oi_1hr_map.get(sym), oi_prevday_map.get(sym),
        )
        if next_oi and next_oi.get("chain"):
            nxt_table = _build_chain_table(sym, next_oi["chain"], {**next_ana, "spot": spot}, {}, next_exp)
            nxt_note  = '<div style="font-size:10px;color:#64748b;padding:4px 0">Refreshed every 15 min</div>'
        else:
            nxt_table = '<div style="color:#64748b;padding:20px;font-size:12px">Next expiry loading...</div>'
            nxt_note  = ""

        vwap_block = build_vwap_html(vwap, sym)
        tech_block = build_technicals_html(td, sym)
        iv_block   = build_iv_html(iv, sym)
        conc_block = build_oi_concentration_html(conc, sym)

        tab_contents[sym] = (
            f'<div class="sym-header">'
            f'<span class="sym-name">{sym}</span>'
            f'<span class="spot-price">{spot:,.1f}</span>'
            f'<span class="pcr-badge">PCR {pcr:.3f}</span>'
            f'<span class="mp-badge">MaxPain {max_pain:,.0f}</span>'
            f'<span class="expiry-badge">{expiry}</span>'
            f'<span class="signal-badge" style="background:{sig_border};color:{sig_col}">{signal}</span>'
            f'</div>'
            f'<div class="signal-card" style="background:{sig_bg};border-color:{sig_border}">'
            f'<div class="signal-main" style="color:{sig_col}">{signal}</div>'
            f'<div class="signal-reasons">{reasons_str}</div>'
            f'</div>'
            f'<div class="setup-card">{best_setup} ({best_score}/10) &nbsp;.&nbsp; {mood}</div>'
            f'<div class="ind-row">{vwap_block}{tech_block}{iv_block}{conc_block}</div>'
            f'<div class="chain-section">'
            f'<div class="tab-bar">'
            f'<button class="tab active" onclick="showChainTab('{sym}','curr',this)">Current: {expiry}</button>'
            f'<button class="tab" onclick="showChainTab('{sym}','next',this)">Next: {next_exp}</button>'
            f'</div>'
            f'<div id="chain-{sym}-curr" class="chain-tab-content">{curr_table}</div>'
            f'<div id="chain-{sym}-next" class="chain-tab-content" style="display:none">{nxt_note}{nxt_table}</div>'
            f'</div>'
        )

    movers_html = _build_movers_html(nifty50_stocks)
    spike_html  = _build_spike_html(delta_map)
    stocks_tab  = (
        '<div class="section-label" style="margin-top:0">Top Movers - Nifty 50</div>'
        + movers_html
        + '<div class="section-label" style="margin-top:16px">OI Spike Alerts</div>'
        + spike_html
    )

    vix_html = ""
    if india_vix > 0:
        vix_col  = "#e74c3c" if india_vix > 20 else "#2ecc71" if india_vix < 14 else "#e67e22"
        vix_html = f'<span class="vix-badge" style="color:{vix_col};font-weight:700">VIX {india_vix:.2f}</span>'

    stop_js = f"var _stopH={stop_h},_stopM={stop_m};"
    refresh_js = (
        stop_js + "
"
        "var _refreshMs=3*60*1000,_startTime=new Date().getTime();
"
        "function _isAfterStop(){var n=new Date();return(n.getHours()>_stopH)||(n.getHours()===_stopH&&n.getMinutes()>=_stopM);}
"
        "function updateCountdown(){
"
        "  if(_isAfterStop()){document.getElementById('countdown-footer').textContent='Market closed';return;}
"
        "  var e=new Date().getTime()-_startTime,r=Math.max(0,Math.ceil((_refreshMs-e)/1000));
"
        "  if(r===0){location.reload();return;}
"
        "  var m=Math.floor(r/60),s=r%60;
"
        "  document.getElementById('countdown-footer').textContent='Next refresh in '+m+'m '+(s<10?'0':'')+s+'s';
"
        "}
"
        "if(!_isAfterStop()){setInterval(updateCountdown,1000);setTimeout(function(){if(!_isAfterStop())location.reload();},_refreshMs);}
"
        "else{document.addEventListener('DOMContentLoaded',function(){var el=document.getElementById('countdown-footer');if(el)el.textContent='Market closed';});}
"
    )
    main_tab_js = (
        "function showMainTab(tabId,btn){
"
        "  document.querySelectorAll('.main-tab-content').forEach(function(el){el.style.display='none';});
"
        "  document.querySelectorAll('.main-tab-btn').forEach(function(el){el.classList.remove('active');});
"
        "  document.getElementById('maintab-'+tabId).style.display='block';
"
        "  btn.classList.add('active');
"
        "}
"
        "function showChainTab(sym,which,btn){
"
        "  ['curr','next'].forEach(function(k){
"
        "    var el=document.getElementById('chain-'+sym+'-'+k);
"
        "    if(el)el.style.display=(k===which)?'block':'none';
"
        "  });
"
        "  btn.closest('.tab-bar').querySelectorAll('.tab').forEach(function(b){b.classList.remove('active');});
"
        "  btn.classList.add('active');
"
        "}
"
    )
    main_tab_css = (
        ".main-tabs{display:flex;gap:0;background:var(--s1);border-bottom:2px solid var(--bd);position:sticky;top:52px;z-index:90;padding:0 24px;}
"
        ".main-tab-btn{background:transparent;border:none;border-bottom:3px solid transparent;color:var(--mt);padding:12px 32px;"
        "font-family:'Inter',sans-serif;font-size:13px;font-weight:700;letter-spacing:1px;cursor:pointer;text-transform:uppercase;"
        "transition:all .2s;margin-bottom:-2px;}
"
        ".main-tab-btn:hover{color:var(--tx);}
"
        ".main-tab-btn.active{color:var(--ac);border-bottom-color:var(--ac);}
"
        ".main-tab-content{display:none;padding:20px 24px;max-width:1500px;margin:0 auto;}
"
    )

    nifty_html     = tab_contents.get("NIFTY", "")
    banknifty_html = tab_contents.get("BANKNIFTY", "")
    pull_time      = ts_str[13:21]
    excel_ts       = last_excel_ts or "---"

    html = (
        "<!DOCTYPE html>
<html lang='en'>
<head>
"
        "<meta charset='UTF-8'>
"
        "<meta name='viewport' content='width=device-width, initial-scale=1.0'>
"
        f"<title>Nifty F&O Dashboard - {ts_str}</title>
"
        "<style>
"
        + _get_base_css() + "
"
        + get_vwap_css() + "
"
        + get_technicals_css() + "
"
        + get_iv_css() + "
"
        + get_oi_concentration_css() + "
"
        + main_tab_css + "
"
        "</style>
</head>
<body>
"
        "<div class='topbar'>
"
        "  <span class='topbar-title'>Nifty F&amp;O Bot v5 - Germany Edition</span>
"
        f"  {vix_html}
"
        f"  <span class='topbar-ts'>{ts_str}</span>
"
        f"  <span class='topbar-meta'>Pull: {pull_time} . Excel: {excel_ts} . Scans: {scan_count}</span>
"
        "</div>

"
        "<div class='main-tabs'>
"
        "  <button class='main-tab-btn active' onclick="showMainTab('nifty',this)">&#9889; NIFTY</button>
"
        "  <button class='main-tab-btn' onclick="showMainTab('banknifty',this)">&#127970; BANKNIFTY</button>
"
        "  <button class='main-tab-btn' onclick="showMainTab('stocks',this)">&#128202; STOCKS</button>
"
        "</div>

"
        "<div id='maintab-nifty' class='main-tab-content' style='display:block'>
"
        f"{nifty_html}
"
        "</div>
"
        "<div id='maintab-banknifty' class='main-tab-content'>
"
        f"{banknifty_html}
"
        "</div>
"
        "<div id='maintab-stocks' class='main-tab-content'>
"
        f"{stocks_tab}
"
        "</div>

"
        "<div class='footer'>
"
        "  Nifty F&amp;O Bot v5 &nbsp;.&nbsp; Germany Edition &nbsp;.&nbsp;
"
        f"  {ts_str} &nbsp;.&nbsp; <span id='countdown-footer'></span>
"
        "</div>
"
        "<script>
"
        + main_tab_js + "
"
        + refresh_js + "
"
        "</script>
</body>
</html>"
    )

    os.makedirs(DATA_FOLDER, exist_ok=True)
    with open(DASH_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    return html



def archive_dashboard():
    """Called at EOD (15:30 IST). Copies dashboard.html to DashboardArchive/dashboard_YYYY-MM-DD.html"""
    if not os.path.exists(DASH_FILE):
        return
    os.makedirs(ARCHIVE_FOLDER, exist_ok=True)
    today = datetime.date.today().strftime("%Y-%m-%d")
    dst   = os.path.join(ARCHIVE_FOLDER, f"dashboard_{today}.html")
    shutil.copy2(DASH_FILE, dst)
    from utils.logger import log
    log.info(f"[Dashboard] Archived to {dst}")


# =============================================================================
#  MOVERS + SPIKES
# =============================================================================

def _build_movers_html(nifty50_stocks: list) -> str:
    if not nifty50_stocks:
        return '<div class="movers-wrap"><div style="padding:10px;color:#4a5568">No stock data</div></div>'
    sorted_stocks = sorted(nifty50_stocks, key=lambda x: x.get("chg_pct", 0), reverse=True)
    gainers = sorted_stocks[:5]
    losers  = sorted_stocks[-5:][::-1]

    def _row(s):
        c   = "#69f0ae" if s.get("chg_pct", 0) >= 0 else "#ff6e6e"
        sgn = "+" if s.get("chg_pct", 0) >= 0 else ""
        return (
            f'<tr><td class="sym-name">{s["symbol"]}</td>'
            f'<td>{s.get("ltp",0):,.1f}</td>'
            f'<td style="color:{c};font-weight:700">{sgn}{s.get("chg_pct",0):.2f}%</td></tr>'
        )

    g_rows = "".join(_row(s) for s in gainers)
    l_rows = "".join(_row(s) for s in losers)
    return (
        f'<div class="movers-wrap">'
        f'<div class="mv-card"><div class="mv-head up">TOP GAINERS</div>'
        f'<table class="mv-table"><tbody>{g_rows}</tbody></table></div>'
        f'<div class="mv-card"><div class="mv-head dn">TOP LOSERS</div>'
        f'<table class="mv-table"><tbody>{l_rows}</tbody></table></div>'
        f'</div>'
    )


def _build_spike_html(delta_map: dict) -> str:
    all_alerts = []
    for sym, delta in delta_map.items():
        if not delta:
            continue
        for tier, a in delta.get("alerts", []):
            all_alerts.append((tier, sym, a))

    if not all_alerts:
        return (
            '<div class="spike-section"><div class="spike-head">'
            '<div class="spike-head-title">SPIKE LOG</div></div>'
            '<table class="spike-table"><thead>'
            '<tr><th>Time</th><th>Symbol</th><th>Tier</th><th>Detail</th></tr>'
            '</thead><tbody>'
            '<tr><td colspan="4" class="nd">No spikes detected this scan</td></tr>'
            '</tbody></table></div>'
        )

    rows = ""
    for tier, sym, a in all_alerts:
        t_col = "#e74c3c" if tier == "EXTREME" else "#e67e22"
        rows += (
            f'<tr>'
            f'<td>{datetime.datetime.now(IST).strftime("%H:%M")}</td>'
            f'<td>{sym}</td>'
            f'<td style="color:{t_col};font-weight:700">{tier}</td>'
            f'<td>{a.get("msg","")}</td>'
            f'</tr>'
        )
    return (
        '<div class="spike-section"><div class="spike-head">'
        '<div class="spike-head-title">SPIKE LOG</div></div>'
        '<table class="spike-table"><thead>'
        '<tr><th>Time</th><th>Symbol</th><th>Tier</th><th>Detail</th></tr>'
        f'</thead><tbody>{rows}</tbody></table></div>'
    )


# =============================================================================
#  BASE CSS
# =============================================================================

def _get_base_css() -> str:
    return """
:root {
  --bg:#0d1117; --s1:#161b22; --s2:#1c2128; --s3:#21262d;
  --tx:#e6edf3; --mt:#7d8590; --bd:#30363d;
  --ac:#f0b429; --green:#3fb950; --red:#f85149;
}
* { box-sizing:border-box; margin:0; padding:0; }
body { background:var(--bg); color:var(--tx); font-family:"Inter","Segoe UI",sans-serif; font-size:13px; }
.topbar { display:flex; align-items:center; gap:12px; padding:8px 18px;
  background:var(--s1); border-bottom:1px solid var(--bd); flex-wrap:wrap; }
.topbar-title { font-weight:700; font-size:13px; color:var(--ac); }
.topbar-ts { color:var(--mt); font-size:11px; margin-left:auto; }
.topbar-meta { color:var(--mt); font-size:10px; }
.vix-badge { font-size:12px; padding:2px 8px; border-radius:6px; background:rgba(255,255,255,0.06); }
.main-content { padding:14px 18px; }
.sym-panel { margin-bottom:22px; }
.sym-header { display:flex; align-items:center; gap:10px; padding:10px 0 8px 0;
  flex-wrap:wrap; border-bottom:1px solid var(--bd); margin-bottom:8px; }
.sym-name { font-size:18px; font-weight:800; color:var(--ac); }
.spot-price { font-size:20px; font-weight:700; font-family:monospace; }
.pcr-badge,.mp-badge,.expiry-badge { font-size:11px; padding:2px 8px; border-radius:6px;
  background:var(--s2); border:1px solid var(--bd); color:var(--mt); }
.signal-badge { font-size:11px; font-weight:700; padding:3px 10px;
  border-radius:8px; border:1px solid; margin-left:auto; }
.signal-card { border:1px solid; border-radius:8px; padding:8px 14px; margin-bottom:6px; }
.signal-main { font-size:14px; font-weight:700; }
.signal-reasons { font-size:11px; color:#94a3b8; margin-top:3px; }
.setup-card { font-size:12px; color:var(--mt); padding:4px 0; margin-bottom:8px; }
.ind-row { display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin-bottom:12px; }
@media(max-width:1200px){.ind-row{grid-template-columns:repeat(2,1fr);}}
@media(max-width:700px){.ind-row{grid-template-columns:1fr;}}
.section-label { font-size:11px; font-weight:700; letter-spacing:1px;
  text-transform:uppercase; color:var(--mt); margin-bottom:8px; padding-bottom:4px;
  border-bottom:1px solid var(--bd); }
.chain-section { background:var(--s1); border:1px solid var(--bd); border-radius:10px; overflow:hidden; margin-bottom:12px; }
.tab-bar { display:flex; gap:4px; padding:8px 12px; background:var(--s2); border-bottom:1px solid var(--bd); }
.tab { background:transparent; border:1px solid var(--bd); color:var(--mt);
  padding:5px 16px; border-radius:6px; cursor:pointer;
  font-family:"Inter",sans-serif; font-size:12px; font-weight:600; transition:all .2s; }
.tab.active { background:var(--ac); color:#000; border-color:var(--ac); font-weight:700; }
.chain-tab-content { overflow-x:auto; }
.ct { width:100%; border-collapse:collapse; font-size:12px; }
.ch th  { padding:8px 10px; font-size:9px; letter-spacing:1px; font-weight:700; text-transform:uppercase; }
.ch2 th { padding:6px 10px; font-size:9px; color:var(--mt); font-weight:700; border-bottom:2px solid var(--bd); }
.ce-head     { background:#1f0a0a; color:#ff6e6e; text-align:center; }
.pe-head     { background:#0a1f0a; color:#69f0ae; text-align:center; }
.strike-head { background:var(--s2); color:var(--ac); text-align:center;
  border-left:1px solid var(--bd); border-right:1px solid var(--bd); }
.ct td { padding:6px 10px; border-bottom:1px solid var(--bg); text-align:center; }
.ct .r { text-align:right; }
.strike-col { font-weight:700; color:var(--ac); background:var(--s2);
  border-left:1px solid var(--bd); border-right:1px solid var(--bd); text-align:center; }
.oi-num { font-weight:600; }
.atm-row td { background:#1a2200 !important; }
.atm-row .strike-col { background:#1a2200 !important; color:#ffeb3b !important; font-weight:700; }
.chain-legend { padding:8px 18px; background:var(--s2); border-top:1px solid var(--bd);
  font-size:10px; color:var(--mt); display:flex; gap:20px; flex-wrap:wrap; }
.bottom-grid { display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:16px; }
.movers-wrap { background:var(--s1); border:1px solid var(--bd); border-radius:10px; overflow:hidden; }
.mv-card { padding:14px 16px; }
.mv-card+.mv-card { border-top:1px solid var(--bd); }
.mv-head { font-size:10px; font-weight:700; letter-spacing:1px; margin-bottom:8px;
  padding-bottom:6px; border-bottom:1px solid var(--bd); }
.mv-head.up { color:var(--green); } .mv-head.dn { color:var(--red); }
.mv-table { width:100%; border-collapse:collapse; font-size:12px; }
.mv-table td { padding:4px 6px; }
.sym-name { font-weight:600; color:var(--tx); }
.spike-section { background:var(--s1); border:1px solid var(--bd); border-radius:10px; overflow:hidden; }
.spike-head { background:var(--s2); border-bottom:1px solid var(--bd); padding:12px 18px; }
.spike-head-title { font-size:12px; font-weight:700; color:var(--tx); letter-spacing:.5px; }
.spike-table { width:100%; border-collapse:collapse; font-size:12px; }
.spike-table th { padding:8px 12px; font-size:9px; letter-spacing:1px; color:var(--mt);
  text-transform:uppercase; border-bottom:1px solid var(--bd); text-align:left; }
.spike-table td { padding:7px 12px; border-bottom:1px solid var(--s2); }
.nd { color:var(--mt); text-align:center; padding:12px; }
.footer { padding:10px 18px; background:var(--s1); border-top:1px solid var(--bd);
  font-size:11px; color:var(--mt); text-align:center; }
"""