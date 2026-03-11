"""
output/excel_writer.py — Excel logging engine for all data sheets.

CHANGES (Phase 1 Rank 1):
  - DATA_DIR now uses updated path from settings (Desktop/Market Data)
  - FUTURES_CANDLES sheet added (OBJ 2 — 3-min OHLCV)
  - OI sheets (_OI) now include per-strike Greeks: Delta, Theta, Gamma, Vega
  - PCR_MAXPAIN sheet now includes VIX column (OBJ 7)
  - FII_DII sheet added (Comment #12)
  - save_futures_candles()  — NEW
  - save_fii_dii()          — NEW
  - save_oi_with_greeks()   — NEW (replaces plain OI append)
  - Atomic write (tmp file + rename) to fix Excel-open permission errors (OBJ 6 partial)
"""

import os
import shutil
import datetime
import tempfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    DATA_DIR, IST,
    COLOUR_HEADER_DARK, COLOUR_HEADER_MID, COLOUR_HEADER_LIGHT,
    COLOUR_GREEN, COLOUR_RED, COLOUR_ORANGE, COLOUR_RED_DARK,
    COLOUR_WHITE, COLOUR_GREY,
)
from utils.logger import log


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED STYLING FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def style_header(cell, text, bg=COLOUR_HEADER_DARK, fg=COLOUR_WHITE, bold=True, size=11):
    cell.value     = text
    cell.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_data(cell, value, bg=None, bold=False, align="center"):
    cell.value     = value
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def append_row(ws, values, row_bg=None):
    ws.append(values)
    row = ws.max_row
    for ci in range(1, len(values) + 1):
        cell = ws.cell(row, ci)
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
        if row_bg:
            cell.fill = PatternFill("solid", start_color=row_bg)
    ws.row_dimensions[row].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
#  WORKBOOK MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

def get_daily_filepath() -> str:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    return str(DATA_DIR / f"MarketData_{date_str}.xlsx")


def load_or_create_workbook(filepath: str):
    """Load existing workbook or create a new one with all sheet definitions."""
    if os.path.exists(filepath):
        try:
            return openpyxl.load_workbook(filepath)
        except Exception as e:
            log.warning(f"[Excel] Could not open {filepath}: {e} — creating fresh")
    return _init_daily_workbook(filepath)


def save_workbook_atomic(wb, filepath: str):
    """
    Save workbook atomically: write to a temp file then rename.
    This avoids the PermissionError when Excel has the file open —
    the rename either succeeds or fails cleanly without corrupting the original.
    """
    dir_  = os.path.dirname(filepath)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=dir_)
    os.close(fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, filepath)
        log.debug(f"[Excel] Saved: {os.path.basename(filepath)}")
    except PermissionError:
        # Excel has the final file open — leave tmp file, log warning
        log.warning(
            f"[Excel] PermissionError saving {os.path.basename(filepath)}. "
            f"Temp file kept at {tmp_path}. Close Excel to unlock."
        )
    except Exception as e:
        log.error(f"[Excel] Save error: {e}")
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def _init_daily_workbook(filepath: str):
    """Create a fresh daily workbook with all required sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_defs = {
        "DASHBOARD": [],
        "NIFTY_OI": [
            "Timestamp", "Expiry", "Spot", "ATM", "Strike",
            "CE_OI", "CE_DayDelta", "CE_Chg%",
            "CE_Delta", "CE_Theta", "CE_Gamma", "CE_Vega",
            "PE_OI", "PE_DayDelta", "PE_Chg%",
            "PE_Delta", "PE_Theta", "PE_Gamma", "PE_Vega",
        ],
        "BANKNIFTY_OI": [
            "Timestamp", "Expiry", "Spot", "ATM", "Strike",
            "CE_OI", "CE_DayDelta", "CE_Chg%",
            "CE_Delta", "CE_Theta", "CE_Gamma", "CE_Vega",
            "PE_OI", "PE_DayDelta", "PE_Chg%",
            "PE_Delta", "PE_Theta", "PE_Gamma", "PE_Vega",
        ],
        "PCR_MAXPAIN": [
            "Timestamp", "Symbol", "Spot", "PCR", "MaxPain", "Spot_vs_MaxPain%",
            "Total_CE_OI", "Total_PE_OI",
            "India_VIX",
            "VWAP", "VWAP_Band1Up", "VWAP_Band1Dn",
            "VWAP_Band2Up", "VWAP_Band2Dn", "VWAP_Weekly_AVWAP", "VWAP_Slope", "VWAP_Position",
        ],
        "OI_DELTA": [
            "Timestamp", "Symbol", "Strike",
            "CE_Prev", "CE_Curr", "CE_3min_Delta", "CE_3min_%",
            "CE_1hr_Delta", "CE_1hr_%",
            "CE_PrevDay_Delta", "CE_PrevDay_%",
            "PE_Prev", "PE_Curr", "PE_3min_Delta", "PE_3min_%",
            "PE_1hr_Delta", "PE_1hr_%",
            "PE_PrevDay_Delta", "PE_PrevDay_%",
            "Mood",
        ],
        "NIFTY50_STOCKS": ["Timestamp", "Symbol", "LTP", "PrevClose", "Change", "Chg%"],
        "SIGNALS": [
            "Timestamp", "Symbol", "Spot", "PCR", "MaxPain",
            "Signal", "Reason1", "Reason2", "Reason3",
        ],
        "SPIKE_LOG": [
            "Timestamp", "Symbol", "Strike", "Side",
            "Prev_OI", "Curr_OI", "3min_Delta", "3min_%", "Tier", "Action",
        ],
        "TECHNICALS": [
            "Timestamp", "Symbol",
            "Supertrend", "Supertrend_Level", "Supertrend_ATR",
            "EMA20", "EMA20_Distance", "EMA20_Distance_Pct", "EMA20_Relation",
            "RSI", "RSI_Zone", "RSI_Divergence",
            "HH_LL", "HH_LL_Detail", "Candles_3m", "As_Of",
        ],
        "FUTURES": [
            "Timestamp", "Symbol", "Contract", "Expiry", "DaysLeft",
            "Futures_LTP", "Spot", "Basis", "Basis%",
            "R1", "R2", "R3", "S1", "S2", "S3",
            "Best_Setup", "CE_Score", "PE_Score", "Straddle_Score", "Strangle_Score",
        ],
        "FUTURES_CANDLES": [
            "Timestamp", "Symbol", "Expiry",
            "Candle_Time", "Open", "High", "Low", "Close", "Volume",
        ],
        "FII_DII": [
            "Date", "FII_Buy_Cr", "FII_Sell_Cr", "FII_Net_Cr",
            "DII_Buy_Cr", "DII_Sell_Cr", "DII_Net_Cr",
            "Source", "Fetched_At",
        ],
    }

    for sheet_name, cols in sheet_defs.items():
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        if cols:
            for ci, col in enumerate(cols, 1):
                style_header(ws.cell(1, ci), col, bg=COLOUR_HEADER_MID)
                ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 2)
            ws.row_dimensions[1].height = 22

    wb.save(filepath)
    log.info(f"[Excel] Created daily workbook: {filepath}")
    return wb


# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET-SPECIFIC SAVE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def save_futures_candles(wb, symbol: str, expiry: str, candles: list):
    """
    Append 3-min futures OHLCV candles to the FUTURES_CANDLES sheet.
    Only appends candles newer than the last logged row (avoids duplicates).
    """
    ws = wb["FUTURES_CANDLES"] if "FUTURES_CANDLES" in wb.sheetnames else wb.create_sheet("FUTURES_CANDLES")
    now_ts = datetime.datetime.now(IST).strftime("%H:%M")

    # Get last logged candle time to avoid duplicates
    last_logged = None
    if ws.max_row > 1:
        last_logged = ws.cell(ws.max_row, 4).value  # Candle_Time column

    for c in candles:
        ctime = c.get("date")
        if ctime is None:
            continue
        # Convert to string for comparison
        ctime_str = ctime.strftime("%H:%M") if hasattr(ctime, "strftime") else str(ctime)[:16]
        if last_logged and ctime_str <= str(last_logged)[:16]:
            continue  # already logged

        append_row(ws, [
            now_ts, symbol, expiry,
            ctime_str,
            round(c.get("open",  0), 2),
            round(c.get("high",  0), 2),
            round(c.get("low",   0), 2),
            round(c.get("close", 0), 2),
            int(c.get("volume",  0)),
        ])


def save_fii_dii(wb, fii_dii: dict):
    """Append FII/DII data to the FII_DII sheet."""
    if not fii_dii or fii_dii.get("error"):
        return
    ws = wb["FII_DII"] if "FII_DII" in wb.sheetnames else wb.create_sheet("FII_DII")
    append_row(ws, [
        fii_dii.get("date", ""),
        fii_dii.get("fii_buy",  0),
        fii_dii.get("fii_sell", 0),
        fii_dii.get("fii_net",  0),
        fii_dii.get("dii_buy",  0),
        fii_dii.get("dii_sell", 0),
        fii_dii.get("dii_net",  0),
        fii_dii.get("source",   ""),
        fii_dii.get("fetched_at", ""),
    ])
    log.info(f"[Excel] FII/DII saved: FII {fii_dii.get('fii_net',0):+,.0f} Cr")


def save_oi_with_greeks(wb, symbol: str, timestamp: str, expiry: str,
                        spot: float, atm: float, chain: dict, iv_per_strike: dict):
    """
    Append ATM-window OI rows with per-strike Greeks to the {symbol}_OI sheet.

    iv_per_strike: {strike: {"CE": {delta, theta, gamma, vega}, "PE": {...}}}
    Pass empty dict if greeks not available — cells will be blank.
    """
    sheet_name = f"{symbol}_OI"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
    if ws is None:
        log.warning(f"[Excel] Sheet {sheet_name} not found")
        return

    for strike in sorted(chain.keys()):
        ce = chain[strike].get("CE", {})
        pe = chain[strike].get("PE", {})
        gk = iv_per_strike.get(strike, {})
        ce_g = gk.get("CE", {})
        pe_g = gk.get("PE", {})

        append_row(ws, [
            timestamp, expiry, round(spot, 2), round(atm, 2), int(strike),
            int(ce.get("openInterest", 0)),
            int(ce.get("changeinOpenInterest", 0)),
            round(ce.get("changeinOpenInterest", 0) / ce.get("openInterest", 1) * 100, 1)
            if ce.get("openInterest", 0) else 0,
            # CE Greeks
            round(ce_g.get("delta", 0), 4),
            round(ce_g.get("theta", 0), 4),
            round(ce_g.get("gamma", 0), 6),
            round(ce_g.get("vega",  0), 4),
            # PE OI
            int(pe.get("openInterest", 0)),
            int(pe.get("changeinOpenInterest", 0)),
            round(pe.get("changeinOpenInterest", 0) / pe.get("openInterest", 1) * 100, 1)
            if pe.get("openInterest", 0) else 0,
            # PE Greeks
            round(pe_g.get("delta", 0), 4),
            round(pe_g.get("theta", 0), 4),
            round(pe_g.get("gamma", 0), 6),
            round(pe_g.get("vega",  0), 4),
        ])
